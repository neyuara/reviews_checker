import requests
from openpyxl import load_workbook
import time
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:134.0) Gecko/20100101 Firefox/134.0'
}
wb = load_workbook("D:\project\wb_reviews\SKU.xlsx")
ws = wb.active
num_of_xlcs = 1
num = 1
def get_info(sku):
    global num
    # запрос для запроса))
    print(num)
    print(sku[:(len(sku)-5)])
    if num<10:
        response = requests.get(f'https://basket-0{num}.wbbasket.ru/vol{sku[:(len(sku)-5)]}/part{sku[:(len(sku)-3)]}/{sku}/info/ru/card.json', headers=headers)
    else:
        response = requests.get(f'https://basket-{num}.wbbasket.ru/vol{sku[:(len(sku)-5)]}/part{sku[:(len(sku)-3)]}/{sku}/info/ru/card.json', headers=headers)
    print(response)
    if response.status_code != 200 :
        num = num +1
        time.sleep(1)
        return get_info(sku)
    response_data_for_imtd = response.json()
    imt_id = response_data_for_imtd["imt_id"]
    print(response_data_for_imtd)
    SKUname = response_data_for_imtd.get("imt_name")
    print(SKUname)
    print('2')
    #запрос отзывов
    response = requests.get(f'https://feedbacks2.wb.ru/feedbacks/v2/{imt_id}', headers=headers)
    response_data = response.json()
    review = response_data.get("feedbacks")[0]
    if review.get("productValuation") !=9:
        print(
                f'''
                Название товара: {SKUname}
                SKU товара: {review.get("nmId", "Неизвестно")}
                Оценка: {review.get("productValuation", "Нет данных")}
                Комментарий: {review.get("text", "Нет данных")}
                Достоинства: {review.get("pros", "Нет данных")}
                Недостатки: {review.get("cons", "Нет данных")}
                '''
            )
    num = 1
def main():
    global  num_of_xlcs, num
    while ws[f"A{num_of_xlcs}"].value:
        sku = ws[f"A{num_of_xlcs}"].value
        print(sku)
        get_info(str(sku))
        num_of_xlcs = num_of_xlcs + 1
        num = 1
if __name__ == "__main__":
    main()
