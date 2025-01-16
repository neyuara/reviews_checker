import requests
import time
from openpyxl import load_workbook

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0',
}
wb = load_workbook("D:\project\wb_reviews\SKU.xlsx")
ws = wb.active
mas_of_reviews = []
SKUname = ''
num_of_xlcs = 1

# Получение имени товара
def get_name(sku):
    global SKUname
    try:
        response = requests.get(f'https://basket-10.wbbasket.ru/vol1545/part154509/{sku}/info/ru/card.json', headers=headers)
        response.raise_for_status()
        data = response.json()
        SKUname = data.get("imt_name", "Неизвестное имя товара")
        print(SKUname)
    except Exception as e:
        print(f"Ошибка при получении имени товара: {e}")
        SKUname = "Неизвестное имя товара"
    print('end of get_name')

def change_response(sku: str):
    global mas_of_reviews
    try:
        response = requests.get(f'https://feedbacks1.wb.ru/feedbacks/v2/{sku}', headers=headers)
    except Exception as e:
        print(f"Ошибка при получении отзывов для SKU {sku}: {e}")
        mas_of_reviews = []
    print('end of change response')

def main():
    global num_of_xlcs
    while ws[f"A{num_of_xlcs}"].value:
        sku = ws[f"A{num_of_xlcs}"].value
        print(f"Обработка SKU: {sku}")
        change_response(sku)
        get_name(sku)
        check_product_reviews()
        num_of_xlcs += 1
        time.sleep(5)
    num_of_xlcs = 1
    time.sleep(60)
    main()

def check_product_reviews():
    if not mas_of_reviews:
        print("Нет отзывов для анализа.")
        return
    for i, review in enumerate(mas_of_reviews):
        print(review)
        if review.get("feedbackCount") == 0 :
            print('нету отзывов')
        elif review.get("productValuation", 0) != 5:
            print(
                f'''
                Название товара: {SKUname}
                SKU товара: {review.get("nmId", "Неизвестно")}
                Оценка: {review.get("productValuation", "Нет данных")}
                Комментарий: {review.get("text", "Нет данных")}
                Достоинства: {review.get("pros", "Нет данных")}
                Недостатки: {review.get("cons", "Нет данных")}
                '''
            )
            print('end of product reviews')
            time.sleep(1)

if __name__ == "__main__":
    main()
